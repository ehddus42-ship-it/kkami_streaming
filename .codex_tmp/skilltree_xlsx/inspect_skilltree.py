import csv
from collections import Counter

from openpyxl import load_workbook

INPUT = r"C:\Users\user\Downloads\20260714_kkami_까미스트리밍데이터테이블_백승오_ver04.xlsx"

workbook = load_workbook(INPUT, read_only=True, data_only=False)

source_sheet = workbook["skilltree"]
source_headers = [cell.value for cell in source_sheet[2]]
source_rows = []
for values in source_sheet.iter_rows(min_row=4, max_col=17, values_only=True):
    if values[0] is None:
        break
    source_rows.append(dict(zip(source_headers, values)))

runtime_path = r"C:\Project\kkami_streaming\Assets\GameKamiStreaming\Resources\GameKamiStreaming\DataTables\skilltree.csv"
with open(runtime_path, encoding="utf-8", newline="") as stream:
    runtime_rows = list(csv.DictReader(stream))

assert len(source_rows) == len(runtime_rows) == 28
runtime_by_id = {int(row["tile_id"]): row for row in runtime_rows}
field_pairs = {
    "upgrade_type": "reinforced_int",
    "upgrade_rank": "upgrade_rank",
    "upgrade_count": "upgrade_count",
    "increase_by": "up_int",
    "follow_amount": "follow_int",
    "watch_amount": "watcher_int",
    "love_amount": "love_int",
    "donation_amount": "donation_int",
    "reddonation_amount": "reddonation_int",
    "subscriber_amount": "subscriber_int",
    "skill_stringkey": "skill_stringkey",
    "skill_name": "skill_name",
}

mismatches = []
for source in source_rows:
    runtime = runtime_by_id[int(source["tile_id"])]
    for source_field, runtime_field in field_pairs.items():
        source_value = source[source_field]
        runtime_value = runtime[runtime_field]
        if source_field in {"skill_stringkey", "skill_name"}:
            if (source_value or "") != runtime_value:
                mismatches.append((source["tile_id"], source_field, source_value, runtime_value))
        else:
            expected = float(source_value or 0)
            actual = float(runtime_value or 0)
            if expected != actual:
                mismatches.append((source["tile_id"], source_field, expected, actual))

print("rows", len(runtime_rows))
print("upgrade_count_distribution", dict(sorted(Counter(int(row["upgrade_count"]) for row in runtime_rows).items())))
print("mismatches", mismatches)
